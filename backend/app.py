from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.dialects.mysql import LONGTEXT
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
import os
import uuid
import json
from datetime import datetime, timedelta
import mysql.connector
from mysql.connector import Error
import hashlib
import random
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import logging
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'your-secret-key-here')
# Build database URI with proper URL encoding for special characters
from urllib.parse import quote_plus
db_username = os.getenv('DB_USERNAME', 'root')
db_password = os.getenv('DB_PASSWORD', 'password')
db_host = os.getenv('DB_HOST', 'localhost')
db_port = os.getenv('DB_PORT', '3306')
db_name = os.getenv('DB_NAME', 'excel_finder_db')

app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql+pymysql://{quote_plus(db_username)}:{quote_plus(db_password)}@{db_host}:{db_port}/{db_name}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.getenv('UPLOAD_FOLDER', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_FILE_SIZE', 16 * 1024 * 1024))  # 16MB max file size

# Create upload directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# Database Models
class User(db.Model):
    __tablename__ = 'users'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    
    # Relationship with Excel files
    excel_files = db.relationship('ExcelFile', backref='user', lazy=True)

class ExcelFile(db.Model):
    __tablename__ = 'excel_files'
    
    id = db.Column(db.Integer, primary_key=True)
    file_id = db.Column(db.String(50), unique=True, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    file_size = db.Column(db.Integer, nullable=False)
    sheets_data = db.Column(LONGTEXT)  # JSON string of sheets data
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

class PasswordReset(db.Model):
    __tablename__ = 'password_resets'
    
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False)
    token = db.Column(db.String(255), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=False)
    is_used = db.Column(db.Boolean, default=False)

# Database initialization
def init_db():
    """Initialize the database and create tables"""
    try:
        with app.app_context():
            db.create_all()
            logger.info("Database tables created successfully")
    except Exception as e:
        logger.error(f"Error creating database tables: {e}")

# Utility functions
def generate_user_id():
    """Generate a unique user ID"""
    return f"USR{str(uuid.uuid4())[:8].upper()}"

def generate_file_id():
    """Generate a unique file ID"""
    return f"FILE{str(uuid.uuid4())[:8].upper()}"

def generate_reset_token():
    """Generate a secure reset token (6-digit OTP)"""
    return ''.join(random.choices(string.digits, k=6))

def send_email(to_email, subject, body):
    """Send email using configured SMTP settings"""
    try:
        # Get SMTP settings from environment variables
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
        smtp_port = int(os.getenv('SMTP_PORT', 587))
        smtp_username = os.getenv('SMTP_USERNAME')
        smtp_password = os.getenv('SMTP_PASSWORD')
        
        # Check if email credentials are configured
        if not smtp_username or not smtp_password or smtp_username == 'your-email@gmail.com':
            logger.warning("Email credentials not configured. Please update .env file with your SMTP settings.")
            return False
        
        msg = MIMEMultipart()
        msg['From'] = smtp_username
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        text = msg.as_string()
        server.sendmail(smtp_username, to_email, text)
        server.quit()
        
        logger.info(f"Email sent successfully to {to_email}")
        return True
    except Exception as e:
        logger.error(f"Error sending email: {e}")
        return False

# Authentication Routes
@app.route('/api/auth/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data or not all(k in data for k in ('email', 'password', 'name')):
            return jsonify({'error': 'Missing required fields'}), 400
        
        email = data['email'].lower().strip()
        password = data['password']
        name = data['name'].strip()
        
        # Check if user already exists
        if User.query.filter_by(email=email).first():
            return jsonify({'error': 'User already exists'}), 409
        
        # Create new user
        user_id = generate_user_id()
        password_hash = generate_password_hash(password)
        
        new_user = User(
            user_id=user_id,
            name=name,
            email=email,
            password_hash=password_hash
        )
        
        db.session.add(new_user)
        db.session.commit()
        
        return jsonify({
            'message': 'User registered successfully',
            'user': {
                'user_id': user_id,
                'name': name,
                'email': email
            }
        }), 201
        
    except Exception as e:
        logger.error(f"Registration error: {e}")
        return jsonify({'error': 'Registration failed'}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        
        if not data or not all(k in data for k in ('email', 'password')):
            return jsonify({'error': 'Missing email or password'}), 400
        
        email = data['email'].lower().strip()
        password = data['password']
        
        # Find user
        user = User.query.filter_by(email=email).first()
        
        if not user or not check_password_hash(user.password_hash, password):
            return jsonify({'error': 'Invalid email or password'}), 401
        
        if not user.is_active:
            return jsonify({'error': 'Account is deactivated'}), 401
        
        return jsonify({
            'message': 'Login successful',
            'user': {
                'user_id': user.user_id,
                'name': user.name,
                'email': user.email
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Login error: {e}")
        return jsonify({'error': 'Login failed'}), 500

@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    try:
        data = request.get_json()
        
        if not data or 'email' not in data:
            return jsonify({'error': 'Email is required'}), 400
        
        email = data['email'].lower().strip()
        
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        token = generate_reset_token()
        expires_at = datetime.utcnow() + timedelta(hours=1)  # Token expires in 1 hour
        
        reset_record = PasswordReset(
            email=email,
            token=token,
            expires_at=expires_at
        )
        
        db.session.add(reset_record)
        db.session.commit()
        
        subject = "Password Reset OTP - Excel Finder"
        body = f"""
Hello {user.name},

You requested a password reset for your Excel Finder account.

Your OTP is: {token}

This OTP will expire in 1 hour.

If you didn't request this password reset, please ignore this email.

Best regards,
Excel Finder Team
        """
        
        email_sent = send_email(email, subject, body)
        
        if email_sent:
            return jsonify({
                'message': 'Password reset OTP sent to your email',
                'email_sent': True
            }), 200
        else:
            # If email fails, still return the token for testing purposes
            logger.warning(f"Email sending failed for {email}, returning token for testing")
            return jsonify({
                'message': 'Password reset token generated (email sending failed)',
                'token': token,  # Remove this in production when email is working
                'email_sent': False
            }), 200
        
    except Exception as e:
        logger.error(f"Forgot password error: {e}")
        return jsonify({'error': 'Failed to process request'}), 500

@app.route('/api/auth/verify-otp', methods=['POST'])
def verify_otp():
    try:
        data = request.get_json()
        
        if not data or not all(k in data for k in ('email', 'otp')):
            return jsonify({'error': 'Email and OTP are required'}), 400
        
        email = data['email'].lower().strip()
        otp = data['otp'].strip()
        
        # Find valid reset record
        reset_record = PasswordReset.query.filter_by(
            email=email,
            token=otp,
            is_used=False
        ).first()
        
        if not reset_record or reset_record.expires_at < datetime.utcnow():
            return jsonify({'error': 'Invalid or expired OTP'}), 400
        
        return jsonify({
            'message': 'OTP verified successfully',
            'token': reset_record.token
        }), 200
        
    except Exception as e:
        logger.error(f"OTP verification error: {e}")
        return jsonify({'error': 'OTP verification failed'}), 500

@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    try:
        data = request.get_json()
        
        if not data or not all(k in data for k in ('token', 'password')):
            return jsonify({'error': 'Token and password are required'}), 400
        
        token = data['token']
        password = data['password']
        
        # Find valid reset record
        reset_record = PasswordReset.query.filter_by(
            token=token, 
            is_used=False
        ).first()
        
        if not reset_record or reset_record.expires_at < datetime.utcnow():
            return jsonify({'error': 'Invalid or expired token'}), 400
        
        # Update user password
        user = User.query.filter_by(email=reset_record.email).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        user.password_hash = generate_password_hash(password)
        reset_record.is_used = True
        
        db.session.commit()
        
        return jsonify({'message': 'Password reset successfully'}), 200
        
    except Exception as e:
        logger.error(f"Reset password error: {e}")
        return jsonify({'error': 'Password reset failed'}), 500

@app.route('/api/upload', methods=['POST'])
def upload_excel():
    """Upload and process Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        user_id = request.form.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'User ID is required'}), 400
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Only Excel files are allowed'}), 400
        
        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404
        

        original_filename = file.filename
        file_id = generate_file_id()
        safe_filename = secure_filename(original_filename)
        if not safe_filename:  
            safe_filename = f"file_{file_id}.xlsx"
        
        unique_filename = f"{file_id}_{safe_filename}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        
        file.save(file_path)
        file_size = os.path.getsize(file_path)
        
        # Process Excel file
        try:
            # Try different engines and approaches for better compatibility
            excel_data = None
            error_messages = []
            
            # Method 1: Try with openpyxl engine (best for .xlsx)
            try:
                excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                logger.info(f"Successfully read Excel file with openpyxl engine")
            except Exception as e1:
                error_messages.append(f"openpyxl: {str(e1)}")
                
                # Method 2: Try with openpyxl and different parameters
                try:
                    excel_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl',
                                            na_values=['', ' ', 'N/A', 'n/a', 'NULL', 'null'])
                    logger.info(f"Successfully read Excel file with openpyxl (with na_values)")
                except Exception as e2:
                    error_messages.append(f"openpyxl_na: {str(e2)}")
                    
                    # Method 3: Try with xlrd engine for .xls files
                    try:
                        excel_data = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
                        logger.info(f"Successfully read Excel file with xlrd engine")
                    except Exception as e3:
                        error_messages.append(f"xlrd: {str(e3)}")
                        
                        # Method 4: Try with calamine engine (if available)
                        try:
                            excel_data = pd.read_excel(file_path, sheet_name=None, engine='calamine')
                            logger.info(f"Successfully read Excel file with calamine engine")
                        except Exception as e4:
                            error_messages.append(f"calamine: {str(e4)}")
                            
                            # Method 5: Try without specifying engine
                            try:
                                excel_data = pd.read_excel(file_path, sheet_name=None)
                                logger.info(f"Successfully read Excel file with default engine")
                            except Exception as e5:
                                error_messages.append(f"default: {str(e5)}")
                                
                                # Method 6: Try reading as binary and converting
                                try:
                                    import openpyxl
                                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                                    excel_data = {}
                                    for sheet_name in workbook.sheetnames:
                                        sheet = workbook[sheet_name]
                                        data = []
                                        for row in sheet.iter_rows(values_only=True):
                                            data.append(row)
                                        if data:
                                            df = pd.DataFrame(data[1:], columns=data[0])
                                            excel_data[sheet_name] = df
                                    logger.info(f"Successfully read Excel file with openpyxl direct method")
                                except Exception as e6:
                                    error_messages.append(f"openpyxl_direct: {str(e6)}")
                                    
                                    # Method 7: Try with different encoding
                                    try:
                                        excel_data = pd.read_excel(file_path, sheet_name=None,
                                                                engine='openpyxl', encoding='utf-8')
                                        logger.info(f"Successfully read Excel file with utf-8 encoding")
                                    except Exception as e7:
                                        error_messages.append(f"utf8_encoding: {str(e7)}")
                                        raise Exception(f"All methods failed. Errors: {'; '.join(error_messages)}")
            
            if excel_data is None or not excel_data:
                raise Exception("Failed to read Excel file with any method")
            
            sheets_data = {}
            
            for sheet_name, df in excel_data.items():
                try:
                    # Convert DataFrame to JSON-serializable format
                    df_cleaned = df.copy()
                    
                    # Handle empty DataFrames
                    if df_cleaned.empty:
                        sheets_data[sheet_name] = []
                        continue
                    
                    # Clean column names (remove special characters)
                    df_cleaned.columns = [str(col).strip() for col in df_cleaned.columns]
                    
                    # Convert all columns to string to avoid JSON serialization issues
                    for col in df_cleaned.columns:
                        try:
                            df_cleaned[col] = df_cleaned[col].astype(str)
                        except Exception as col_error:
                            logger.warning(f"Error converting column {col}: {col_error}")
                            df_cleaned[col] = df_cleaned[col].apply(lambda x: str(x) if pd.notna(x) else '')
                    
                    # Replace various forms of NaN and empty values
                    df_cleaned = df_cleaned.replace(['nan', 'NaN', 'NAN', 'None', 'NULL', 'null', 'N/A', 'n/a'], '')
                    df_cleaned = df_cleaned.fillna('')
                    
                    # Remove completely empty rows
                    df_cleaned = df_cleaned.dropna(how='all')
                    
                    # Convert to records
                    records = df_cleaned.to_dict('records')
                    sheets_data[sheet_name] = records
                    
                except Exception as sheet_error:
                    logger.error(f"Error processing sheet {sheet_name}: {sheet_error}")
                    # If a sheet fails, create an empty sheet
                    sheets_data[sheet_name] = []
            
            # Save to database
            excel_file = ExcelFile(
                file_id=file_id,
                user_id=user.id,
                filename=unique_filename,
                original_filename=original_filename,
                file_path=file_path,
                file_size=file_size,
                sheets_data=json.dumps(sheets_data)
            )
            
            db.session.add(excel_file)
            db.session.commit()
            
            logger.info(f"Successfully processed and saved Excel file: {original_filename}")
            
            return jsonify({
                'message': 'File uploaded successfully',
                'file_id': file_id,
                'filename': original_filename,
                'sheets': list(sheets_data.keys()),
                'sheets_data': sheets_data
            }), 201
            
        except Exception as e:
            # Clean up file if processing fails
            if os.path.exists(file_path):
                os.remove(file_path)
            logger.error(f"Excel processing error for {original_filename}: {e}")
            return jsonify({
                'error': f'Failed to process Excel file: {str(e)}',
                'details': 'Please ensure the file is a valid Excel file (.xlsx or .xls) and not corrupted. The file may contain unsupported formatting or be password protected.'
            }), 500
            
    except Exception as e:
        logger.error(f"Upload error: {e}")
        return jsonify({'error': 'Upload failed'}), 500

@app.route('/api/files/<user_id>', methods=['GET'])
def get_user_files(user_id):
    """Get all files uploaded by a user"""
    try:
        user = User.query.filter_by(user_id=user_id).first()
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        files = ExcelFile.query.filter_by(user_id=user.id, is_active=True).all()
        
        files_data = []
        for file in files:
            files_data.append({
                'file_id': file.file_id,
                'filename': file.original_filename,
                'uploaded_at': file.uploaded_at.isoformat(),
                'file_size': file.file_size
            })
        
        return jsonify({'files': files_data}), 200
        
    except Exception as e:
        logger.error(f"Get files error: {e}")
        return jsonify({'error': 'Failed to retrieve files'}), 500

@app.route('/api/files/<file_id>/data', methods=['GET'])
def get_file_data(file_id):
    """Get Excel file data by file ID"""
    try:
        file = ExcelFile.query.filter_by(file_id=file_id, is_active=True).first()
        if not file:
            return jsonify({'error': 'File not found'}), 404
        
        sheets_data = json.loads(file.sheets_data)
        
        return jsonify({
            'file_id': file.file_id,
            'filename': file.original_filename,
            'sheets': list(sheets_data.keys()),
            'sheets_data': sheets_data
        }), 200
        
    except Exception as e:
        logger.error(f"Get file data error: {e}")
        return jsonify({'error': 'Failed to retrieve file data'}), 500

@app.route('/api/files/<file_id>', methods=['DELETE'])
def delete_file(file_id):
    """Delete an Excel file"""
    try:
        file = ExcelFile.query.filter_by(file_id=file_id).first()
        if not file:
            return jsonify({'error': 'File not found'}), 404
        
        # Delete physical file
        if os.path.exists(file.file_path):
            os.remove(file.file_path)
        
        # Mark as inactive in database
        file.is_active = False
        db.session.commit()
        
        return jsonify({'message': 'File deleted successfully'}), 200
        
    except Exception as e:
        logger.error(f"Delete file error: {e}")
        return jsonify({'error': 'Failed to delete file'}), 500

# Health check endpoint
@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()}), 200

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
