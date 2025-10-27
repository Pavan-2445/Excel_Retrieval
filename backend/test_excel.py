import pandas as pd
import openpyxl
import os
import sys
from pathlib import Path

def test_excel_file(file_path):
    print(f"Testing file: {file_path}")
    print(f"File size: {os.path.getsize(file_path)} bytes")
    print("-" * 50)
    
    methods = [
        ("pandas + openpyxl", lambda: pd.read_excel(file_path, sheet_name=None, engine='openpyxl')),
        ("pandas + openpyxl (na_values)", lambda: pd.read_excel(file_path, sheet_name=None, engine='openpyxl', na_values=['', ' ', 'N/A', 'n/a', 'NULL', 'null'])),
        ("pandas + xlrd", lambda: pd.read_excel(file_path, sheet_name=None, engine='xlrd')),
        ("pandas default", lambda: pd.read_excel(file_path, sheet_name=None)),
        ("openpyxl direct", lambda: test_openpyxl_direct(file_path)),
    ]
    
    for method_name, method_func in methods:
        try:
            print(f"Testing {method_name}...")
            result = method_func()
            
            if isinstance(result, dict):
                print(f"SUCCESS: Found {len(result)} sheets")
                for sheet_name, df in result.items():
                    print(f"   Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
            else:
                print(f"SUCCESS: {type(result)}")
                
        except Exception as e:
            print(f"FAILED: {str(e)}")
        print()

def test_openpyxl_direct(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    result = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        if data:
            df = pd.DataFrame(data[1:], columns=data[0])
            result[sheet_name] = df
        else:
            result[sheet_name] = pd.DataFrame()
    
    return result

def main():
    if len(sys.argv) != 2:
        print("Usage: python test_excel.py <excel_file_path>")
        sys.exit(1)
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        sys.exit(1)
    if not file_path.lower().endswith(('.xlsx', '.xls')):
        print("Please provide an Excel file (.xlsx or .xls)")
        sys.exit(1)
    test_excel_file(file_path)

if __name__ == "__main__":
    main()
